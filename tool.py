from openpyxl  import load_workbook
import re
def main(name,cs):
    wb = load_workbook('static/{}'.format(name))
    if len(wb.sheetnames) == 0:
        return 'err0'
    sheet = wb[wb.sheetnames[0]]
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    for i in range(1,maxColumn+1):
        if sheet.cell(1,i).value == '便签':
            bqIndex = i
    for i in range(2,maxRow+1):
        value = sheet.cell(i,bqIndex).value
        dds = re.findall(r'订单号:(\d+)',value)
        cgList = re.findall(r'{}.*?-(-?\d+\.\d+)?'.format(cs),value)
        print(cgList)
        sheet.cell(row=i, column=bqIndex+1,value=cs)
        for j in range(bqIndex+2,bqIndex+2+len(cgList)):
            sheet.cell(row=i, column=j,value=cgList[j-(bqIndex+2)])
            sheet.cell(row=i, column=j+1,value=','.join(dds))
            # print(cgList[j-(bqIndex+1)])
    newName = 'update-{}'.format(name)
    wb.save('static/{}'.format(newName))
    return newName


if __name__ == '__main__':
    main('16686762065287 .xlsx',"10.1审核")