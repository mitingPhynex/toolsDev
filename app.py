import time
from decimal import Decimal
import pandas as pd
from openpyxl  import load_workbook
import re
from flask import Flask, render_template, request, send_from_directory,jsonify
from datetime import datetime

from bq import bq
from bq1 import bq1
from bq2 import bq2
from generateAndsplit import generateMultiColor,generate,generate2,generateSportshirt,generateNew,generateStock,generateGuatan,generateStockGT,generateGuatanWithColor,generateStockGTWithColor
from generateAndsplit import generateGuatan2, generateGuatanWithColor2, generateMultiCloth, generateStockMulti, generateYulianWithColor, generateTikTokCloth, generateSanjiantaoWithColor
from fy import fy
from pp import pp, getRule

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'upload/'

@app.route('/')
def upload_file():
    return render_template('index.html')

@app.route('/tool')
def toolHtml():
    return render_template('tool.html')

@app.route('/uploaderPP',methods=['GET','POST'])
def uploaderPP():
    if request.method == 'POST':
        f = request.files['file']
        rule = request.files['rule']
        name = '{}{}'.format(int(time.time() * 10000), '.xlsx')
        rule.save('./static/{}'.format(name))
        r = getRule('./static/{}'.format(name))
        str(r).replace("\xa0"," ")
        #print(str(r).replace("\xa0"," "))
        name = '{}{}'.format(int(time.time() * 10000), '.xlsx')
        f.save('./static/{}'.format(name))
        newName = pp('./static/{}'.format(name),name,r)
        return jsonify({"name":newName,'code':200})

@app.route('/uploaderSC',methods=['GET','POST'])
def uploaderSC():
    if request.method == 'POST':
        f = request.files['file']
        num = int(request.form.get("num"))
        name = '{}{}'.format(int(time.time() * 10000), '.xlsx')
        f.save('./static/{}'.format(name))
        zipName = bq('./static/{}'.format(name),name,num)
        return jsonify({"name":zipName,'code':200})

#复制生成和分割
@app.route('/uploaderSC1',methods=['GET','POST'])
def uploaderSC1():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        #按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generate('./static/{}'.format(name),name,num)
        return jsonify({"name":zipName,'code':200})

#复制生成和分割，添加视频和视频封面
@app.route('/uploaderSC6',methods=['GET','POST'])
def uploaderSC6():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        #按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateNew('./static/{}'.format(name),name,num)
        return jsonify({"name":zipName,'code':200})

#多配色复制生成和分割旧版
# @app.route('/uploaderSC2',methods=['GET','POST'])
# def uploaderSC2():
#     if request.method == 'POST':
#         file = request.files['file']
#         fileName = file.filename.split(".")[0]
#         num = int(request.form.get("num"))
#         #按照原文件名+时间戳的格式保存上传文件
#         name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
#         file.save('./static/{}'.format(name))
#         zipName = generateMultiColor('./static/{}'.format(name),name,num)
#         return jsonify({"name":zipName,'code':200})

#多配色复制生成和分割 
@app.route('/uploaderSC3',methods=['GET','POST'])
def uploaderSC3():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        #按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateMultiColor('./static/{}'.format(name),name,num)
        return jsonify({"name":zipName,'code':200})
    
#复制生成和分割
@app.route('/uploaderSC4',methods=['GET','POST'])
def uploaderSC4():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        #按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generate2('./static/{}'.format(name),name,num)
        return jsonify({"name":zipName,'code':200})
    
# #复制生成和分割连帽卫衣专用
# @app.route('/uploaderSC5',methods=['GET','POST'])
# def uploaderSC5():
#     if request.method == 'POST':
#         file = request.files['file']
#         fileName = file.filename.split(".")[0]
#         num = int(request.form.get("num"))
#         #按照原文件名+时间戳的格式保存上传文件
#         name = '{}{}'.format(str(fileName)+getCurrentDateTimeFormatted(), '.xlsx')
#         file.save('./static/{}'.format(name))
#         zipName = generateSportshirt('./static/{}'.format(name),name,num)
#         return jsonify({"name":zipName,'code':200})
    
#复制生成和分割通用版
@app.route('/uploaderSC5', methods=['GET', 'POST'])
def uploaderSC5():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        russian_sizes = request.form.get("russianSizes").split(",")
        manufacturer_sizes = request.form.get("manufacturerSizes").split(",")
        files_symbol = request.form.get("files_symbol").split("\n")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateSportshirt('./static/{}'.format(name), name, num, russian_sizes, manufacturer_sizes, files_symbol)
        return jsonify({"name": zipName, 'code': 200})
 
#生成库存数据
@app.route('/uploaderSC7', methods=['GET', 'POST'])
def uploaderSC7():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        russian_sizes = request.form.get("russianSizes").split(",")
        manufacturer_sizes = request.form.get("manufacturerSizes").split(",")
        ck_name = request.form.get("ck_name").split("\n")
        files_symbol = request.form.get("files_symbol").split("\n")
        stock_count = request.form.get("stock_count")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateStock('./static/{}'.format(name), name, num, russian_sizes, manufacturer_sizes, ck_name, stock_count, files_symbol)
        return jsonify({"name": zipName, 'code': 200})
       
#多配色复制生成和分割通用版
@app.route('/uploaderSC15', methods=['GET', 'POST'])
def uploaderSC15():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        russian_sizes = request.form.get("russianSizes").split(",")
        manufacturer_sizes = request.form.get("manufacturerSizes").split(",")
        files_symbol = request.form.get("files_symbol").split("\n")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateMultiCloth('./static/{}'.format(name), name, num, russian_sizes, manufacturer_sizes, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#多配色生成库存数据
@app.route('/uploaderSC16', methods=['GET', 'POST'])
def uploaderSC16():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        russian_sizes = request.form.get("russianSizes").split(",")
        manufacturer_sizes = request.form.get("manufacturerSizes").split(",")
        ck_name = request.form.get("ck_name").split("\n")
        files_symbol = request.form.get("files_symbol").split("\n")
        stock_count = request.form.get("stock_count")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateStockMulti('./static/{}'.format(name), name, num, russian_sizes, manufacturer_sizes, ck_name, stock_count, files_symbol)
        return jsonify({"name": zipName, 'code': 200})


#挂毯
@app.route('/uploaderSC8', methods=['GET', 'POST'])
def uploaderSC8():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateGuatan('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#挂毯放挂毯类目
@app.route('/uploaderSC13', methods=['GET', 'POST'])
def uploaderSC13():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateGuatan2('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})
    
#挂毯生成库存数据
@app.route('/uploaderSC9', methods=['GET', 'POST'])
def uploaderSC9():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        ck_name = request.form.get("ck_name").split("\n")
        id_symbol = request.form.get("idSymbol").split(",")
        files_symbol = request.form.get("files_symbol").split("\n")
        stock_count = request.form.get("stock_count")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateStockGT('./static/{}'.format(name), name, num, id_symbol, ck_name, stock_count, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#多配色挂毯
@app.route('/uploaderSC11', methods=['GET', 'POST'])
def uploaderSC11():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateGuatanWithColor('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#多配色挂毯放挂毯类目
@app.route('/uploaderSC14', methods=['GET', 'POST'])
def uploaderSC14():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateGuatanWithColor2('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})
    
#多配色挂毯、浴帘生成库存数据
@app.route('/uploaderSC12', methods=['GET', 'POST'])
def uploaderSC12():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        ck_name = request.form.get("ck_name").split("\n")
        id_symbol = request.form.get("idSymbol").split(",")
        files_symbol = request.form.get("files_symbol").split("\n")
        stock_count = request.form.get("stock_count")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateStockGTWithColor('./static/{}'.format(name), name, num, id_symbol, ck_name, stock_count, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#多配色浴帘
@app.route('/uploaderSC17', methods=['GET', 'POST'])
def uploaderSC17():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateYulianWithColor('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#多配色三件套
@app.route('/uploaderSC41', methods=['GET', 'POST'])
def uploaderSC41():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        num = int(request.form.get("num"))
        id_symbol = request.form.get("idSymbol").split(",")
        price = request.form.get("discountPrice").split(",")
        price_before_discount = request.form.get("originalPrice").split(",")
        length_of_sku = request.form.get("lengthOfSku").split(",")
        width_of_sku = request.form.get("widthOfSku").split(",")
        files_symbol = request.form.get("files_symbol").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateSanjiantaoWithColor('./static/{}'.format(name), name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol)
        return jsonify({"name": zipName, 'code': 200})

#翻译
# @app.route('/uploaderFY',methods=['GET','POST'])
# def uploaderFY():
#     if request.method == 'POST':
#         f = request.files['file']
#         name = '{}{}'.format(int(time.time() * 10000), '.xlsx')
#         f.save('./static/{}'.format(name))
#         newName = fy('./static/{}'.format(name),name)
#         return jsonify({"name":newName,'code':200})



@app.route('/uploader',methods=['GET','POST'])
def uploader():
    if request.method == 'POST':
        f = request.files['file']
        cs = request.values.get('cs')
        name = '{}{}'.format(int(time.time()*10000),'.xlsx')
        f.save('./static/{}'.format(name))
        newName,zj = star(name,cs)
        return jsonify({"name":newName,'code':200,"zj":zj})


@app.route("/download/<filename>", methods=['GET'])
def download_file(filename):
    return send_from_directory('static', filename, as_attachment=True)



def getCurrentDateTimeFormatted(format_str="%Y-%m-%d %H%M%S"):
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime(format_str)
    return formatted_datetime

#TikTok自动生成服装尺码
@app.route('/uploaderSC31', methods=['GET', 'POST'])
def uploaderSC31():
    if request.method == 'POST':
        file = request.files['file']
        fileName = file.filename.split(".")[0]
        manufacturer_sizes = request.form.get("manufacturerSizes").split(",")

        # 按照原文件名+时间戳的格式保存上传文件
        name = '{}{}'.format(str(fileName) + getCurrentDateTimeFormatted(), '.xlsx')
        file.save('./static/{}'.format(name))
        zipName = generateTikTokCloth('./static/{}'.format(name), name, manufacturer_sizes)
        return jsonify({"name": zipName, 'code': 200})


def star(name,cs):
    wb = load_workbook('static/{}'.format(name))
    if len(wb.sheetnames) == 0:
        return 'err0'
    zj = 0
    sheet = wb[wb.sheetnames[0]]
    maxRow = sheet.max_row
    maxColumn = sheet.max_column
    for i in range(1,maxColumn+1):
        if sheet.cell(1,i).value == '便签':
            bqIndex = i
    for i in range(2,maxRow+1):
        value = str(sheet.cell(i,bqIndex).value)
        dds = re.findall(r'{}.*?订单号[:|：|\s](\d+)'.format(cs),value)
        cgList = re.findall(r'{}[\s\S]*?\d+[\-A-Za-z0-9\s]*-(\d+\.\d+|\d+)'.format(cs),value)
        NEWlIST = []
        for q in cgList:
            if float(q)<10000000:
                NEWlIST.append(q)
                zj = zj + Decimal(q)

        cgList = NEWlIST
        sheet.cell(row=i, column=bqIndex+1,value=cs)
        for j in range(bqIndex+2,bqIndex+2+len(cgList)):
            sheet.cell(row=i, column=j,value=cgList[j-(bqIndex+2)])
            sheet.cell(row=i, column=j+1,value=','.join(dds))
            # print(cgList[j-(bqIndex+1)])
    newName = 'update-{}'.format(name)
    wb.save('static/{}'.format(newName))
    return newName,zj


if __name__ == '__main__':
   app.run(debug=True,host='0.0.0.0',port=5002,threaded=True)