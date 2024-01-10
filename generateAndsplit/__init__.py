import pandas as pd
from openpyxl import load_workbook
import zipfile
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os
import logging
logging.basicConfig(level=logging.INFO)

def cut(obj, sec):
    if sec == 0:
        return [obj]
    return [obj[i:i+sec] for i in range(0,len(obj),sec)]

# 输入最初的文件
def generate(input,name,num):
    # 新文件名数组
    newNameLists = []
    # 生成文件的模板
    template_path = './generateAndsplit/mb_Supplier template.xlsx'
    zipName = name.split(".")[0]+"zip_over.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    #表示用第几行作为表头，默认header=0，即默认第一行为表头
    # df = pd.read_excel('input.xlsx', sheet_name="Vendor template",header=1)
    df = pd.read_excel(input, sheet_name="Supplier template",header=1)
    russian_sizes = ["46", "48", "50", "52", "54", "56"]
    manufacturer_sizes = ['L', 'XL', '2XL', '3XL', '4XL', '5XL']

    # 将每一行分成六份，并将它们在一起
    new_rows = []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(6):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    # 写入数据
    for index,li in enumerate(cut(new_df,num)):
        template_wb = load_workbook(template_path)
        worksheet = template_wb.active
        newName = name.split(".")[0] + '_'+ str(index+1) +"_over.xlsx"
        for row in dataframe_to_rows(li, index=False, header=False):
            worksheet.append(row)
        for cell in worksheet['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        # new_df.to_excel("../static/{}".format(newName), index=False,header=None)
    zipFile.close()
    return zipName

# 添加视频生成和拆分
def generateNew(input, name, num):
    newNameLists = []
    template_path = './generateAndsplit/mb_Supplier template.xlsx'
    zipName = name.split(".")[0]+"zip_over.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    
    df = pd.read_excel(input, sheet_name="Supplier template", header=1)
    
    #视频封面直接获取第一个，之后所有的sku都设置一样的视频
    df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
    video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"]

    #视频直接获取第一行，之后所有的sku都设置一样的视频
    df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
    video_name = df_video.iloc[1]["Ozone.Video: Name"]
    video_url = df_video.iloc[1]["Ozon.Video: URL"]
    video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"]
 
    russian_sizes = ["46", "48", "50", "52", "54", "56"]
    manufacturer_sizes = ['L', 'XL', '2XL', '3XL', '4XL', '5XL']

    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(6):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name
            df_video.loc[i, "Ozon.Video: URL"] = video_url
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video
            row = df_video.iloc[i].tolist()
            new_video.append(row)


    columns = df.columns.tolist()
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Supplier template" sheet
        worksheet_supplier = template_wb["Supplier template"]
        newName = name.split(".")[0] + '_' + str(index + 1) + "_over.xlsx"
        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # Write to "Ozone.Video cover" sheet
        worksheet_video_cover = template_wb["Ozone.Video cover"]
        for row in dataframe_to_rows(li_video_cover, index=False, header=False):
            worksheet_video_cover.append(row)

        # Write to "Ozone.Video" sheet
        worksheet_video = template_wb["Ozone.Video"]
        for row in dataframe_to_rows(li_video, index=False, header=False):
            worksheet_video.append(row)

        # Additional operations for "Supplier template" sheet
        for cell in worksheet_supplier['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet_supplier['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass

        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)

    zipFile.close()
    return zipName


# 韩语服饰短袖T恤，7个sku
def generate2(input,name,num):
    newNameLists = []
    template_path = './generateAndsplit/mb_Supplier template.xlsx'
    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    # df = pd.read_excel('input.xlsx', sheet_name="Vendor template",header=1)
    df = pd.read_excel(input, sheet_name="Supplier template",header=1)
    russian_sizes = ["46", "48", "50", "52", "54", "56", "58"]
    manufacturer_sizes = ['S', 'M', 'L', 'XL', '2XL', '3XL', '4XL']

    # 将每一行分成六份，并将它们在一起
    new_rows = []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(7):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    # 写入数据
    for index,li in enumerate(cut(new_df,num)):
        template_wb = load_workbook(template_path)
        worksheet = template_wb.active
        newName = name.split(".")[0] + '_'+ str(index+1) +"_商品.xlsx"
        for row in dataframe_to_rows(li, index=False, header=False):
            worksheet.append(row)
        for cell in worksheet['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        # new_df.to_excel("../static/{}".format(newName), index=False,header=None)
    zipFile.close()
    return zipName

#多配色服饰通用版
def generateMultiCloth(input, name, num, russian_sizes, manufacturer_sizes, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(russian_sizes)):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            df.loc[i, "Article code*"] = df.iloc[i]["Merge into One PDP*"] + " " + df.iloc[i]["Product color*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Article code*"] = df.iloc[i]["Merge into One PDP*"] + " " + df.iloc[i]["Product color*"] + " " + manufacturer_sizes[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Article code*"] = df.iloc[i]["Merge into One PDP*"] + " " + df.iloc[i]["Product color*"] + " " + manufacturer_sizes[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    logging.info(f"files_symbol{files_symbol}")

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Supplier template" sheet
        worksheet_supplier = template_wb["Template"]
        # newName = name.split(".")[0] + '_' + str(index + 1) + "_商品.xlsx"

        # logging.info(f"files_symbol………………{files_symbol}")
        # logging.info(f"files_symbol[index]………………{files_symbol[index]}")
        # if files_symbol:
        #     newName = f"{name.split('.')[0]}_{str(index + 1)}_sku_{files_symbol[index]}.xlsx"
        # else:
        #     newName = f"{name.split('.')[0]}_{str(index + 1)}_sku.xlsx"
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)

        # Additional operations for "Supplier template" sheet
        for cell in worksheet_supplier['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet_supplier['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass

        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)

    zipFile.close()
    return zipName

#多配色服饰生成库存
def generateStockMulti(input, name, num, russian_sizes, manufacturer_sizes, ck_name, stock_count, files_symbol):
    logging.info("Starting generateStock function...")

    newNameLists = []
    template_path = './generateAndsplit/stock-update-template-cn.xlsx'
    zipName = name.split(".")[0] + "zip_库存.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')

    logging.info("Reading Excel file...")
    df = pd.read_excel(input, sheet_name="Template", header=1)

    # 生成id列表
    product_id_list = []
    logging.info("Generating product ID list...")
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(manufacturer_sizes)):
            product_id = df.iloc[i]["Merge into One PDP*"] + " " + df.iloc[i]["Product color*"] + " " + manufacturer_sizes[j]
            product_id_list.append(product_id)

    # 当num为0时，不分割product_id_list
    if num == 0:
        product_id_chunks = [product_id_list]
    else:
        product_id_chunks = [product_id_list[i:i + num] for i in range(0, len(product_id_list), num)]

    # 检查分割后数组的长度与ck_name的长度是否一致
    if len(product_id_chunks) != len(ck_name) and num != 0:
        error_message = 'product_id_list的分割长度与ck_name的长度不一致'
        logging.error(error_message)
        return {'error': error_message}

    # 写入数据
    logging.info("Writing data...")
    for index, chunk in enumerate(product_id_chunks):
        logging.info(f"Processing chunk {index + 1}/{len(product_id_chunks)}...")
        template_wb = load_workbook(template_path)

        # 写入“Warehouse stock”表单
        worksheet_warehouse = template_wb["仓库库存"]

        # 确定仓库名称
        warehouse_name = ck_name[0] if num == 0 else ck_name[index]

        # 确定开始行
        row_to_start = 2
        
        # 写入整行数据
        for product_index, product_id in enumerate(chunk):
            row_number = product_index + row_to_start
            row_data = (warehouse_name, product_id, '', int(stock_count))
            for col, value in enumerate(row_data, start=1):
                worksheet_warehouse.cell(row=row_number, column=col, value=value)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_库存_' + str(index + 1) + '_' + symbol + ".xlsx"

        # 保存当前工作簿
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")

    zipFile.close()
    logging.info(f"Zip file created: {zipName}")

    return zipName

# 服饰通用版
def generateSportshirt(input, name, num, russian_sizes, manufacturer_sizes, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Supplier template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Supplier template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(russian_sizes)):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    logging.info(f"files_symbol{files_symbol}")

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Supplier template" sheet
        worksheet_supplier = template_wb["Supplier template"]
        # newName = name.split(".")[0] + '_' + str(index + 1) + "_商品.xlsx"

        # logging.info(f"files_symbol………………{files_symbol}")
        # logging.info(f"files_symbol[index]………………{files_symbol[index]}")
        # if files_symbol:
        #     newName = f"{name.split('.')[0]}_{str(index + 1)}_sku_{files_symbol[index]}.xlsx"
        # else:
        #     newName = f"{name.split('.')[0]}_{str(index + 1)}_sku.xlsx"
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)

        # Additional operations for "Supplier template" sheet
        for cell in worksheet_supplier['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet_supplier['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass

        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)

    zipFile.close()
    return zipName

#服饰生成库存
def generateStock(input, name, num, russian_sizes, manufacturer_sizes, ck_name, stock_count, files_symbol):
    logging.info("Starting generateStock function...")

    newNameLists = []
    template_path = './generateAndsplit/stock-update-template-cn.xlsx'
    zipName = name.split(".")[0] + "zip_库存.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')

    logging.info("Reading Excel file...")
    df = pd.read_excel(input, sheet_name="Supplier template", header=1)

    # 生成id列表
    product_id_list = []
    logging.info("Generating product ID list...")
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(manufacturer_sizes)):
            product_id = df.iloc[i]["Merge on One PDP*"] + " " + manufacturer_sizes[j]
            product_id_list.append(product_id)

    # 当num为0时，不分割product_id_list
    if num == 0:
        product_id_chunks = [product_id_list]
    else:
        product_id_chunks = [product_id_list[i:i + num] for i in range(0, len(product_id_list), num)]

    # 检查分割后数组的长度与ck_name的长度是否一致
    if len(product_id_chunks) != len(ck_name) and num != 0:
        error_message = 'product_id_list的分割长度与ck_name的长度不一致'
        logging.error(error_message)
        return {'error': error_message}

    # 写入数据
    logging.info("Writing data...")
    for index, chunk in enumerate(product_id_chunks):
        logging.info(f"Processing chunk {index + 1}/{len(product_id_chunks)}...")
        template_wb = load_workbook(template_path)

        # 写入“Warehouse stock”表单
        worksheet_warehouse = template_wb["仓库库存"]

        # 确定仓库名称
        warehouse_name = ck_name[0] if num == 0 else ck_name[index]

        # 确定开始行
        row_to_start = 2
        
        # 写入整行数据
        for product_index, product_id in enumerate(chunk):
            row_number = product_index + row_to_start
            row_data = (warehouse_name, product_id, '', int(stock_count))
            for col, value in enumerate(row_data, start=1):
                worksheet_warehouse.cell(row=row_number, column=col, value=value)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_库存_' + str(index + 1) + '_' + symbol + ".xlsx"

        # 保存当前工作簿
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")

    zipFile.close()
    logging.info(f"Zip file created: {zipName}")

    return zipName

# 挂毯通用版
def generateGuatan(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Length of the Larger Side, cm'] = length_of_sku[j]
            df.loc[i, 'Length of the smaller side, cm'] = width_of_sku[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)

        # Additional operations for "Template" sheet
        # for cell in worksheet_supplier['U']:
        #     if cell.value is not None:
        #         try:
        #             cell.value = int(cell.value)
        #         except:
        #             pass
        # for cell in worksheet_supplier['AF']:
        #     if cell.value is not None:
        #         try:
        #             cell.value = int(cell.value)
        #         except:
        #             pass
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")
    zipFile.close()
    return zipName

# 挂毯放挂毯类目
def generateGuatan2(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Length of the Larger Side, cm*'] = length_of_sku[j]
            df.loc[i, 'Length of the smaller side, cm'] = width_of_sku[j]
            df.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Product ID*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)

        # Additional operations for "Template" sheet
        # for cell in worksheet_supplier['U']:
        #     if cell.value is not None:
        #         try:
        #             cell.value = int(cell.value)
        #         except:
        #             pass
        # for cell in worksheet_supplier['AF']:
        #     if cell.value is not None:
        #         try:
        #             cell.value = int(cell.value)
        #         except:
        #             pass
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")
    zipFile.close()
    return zipName

#挂毯生成库存
def generateStockGT(input, name, num, id_symbol, ck_name, stock_count, files_symbol):
    logging.info("Starting generateStock function...")

    newNameLists = []
    template_path = './generateAndsplit/stock-update-template-cn.xlsx'
    zipName = name.split(".")[0] + "zip_库存.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')

    logging.info("Reading Excel file...")
    df = pd.read_excel(input, sheet_name="Template", header=1)

    # 生成id列表
    product_id_list = []
    logging.info("Generating product ID list...")
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            product_id = df.iloc[i]["Model Name (to combine products into one PDP)*"] + id_symbol[j]
            product_id_list.append(product_id)

    # 当num为0时，不分割product_id_list
    if num == 0:
        product_id_chunks = [product_id_list]
    else:
        product_id_chunks = [product_id_list[i:i + num] for i in range(0, len(product_id_list), num)]

    # 检查分割后数组的长度与ck_name的长度是否一致
    if len(product_id_chunks) != len(ck_name) and num != 0:
        error_message = 'product_id_list的分割长度与ck_name的长度不一致'
        logging.error(error_message)
        return {'error': error_message}

    # 写入数据
    logging.info("Writing data...")
    for index, chunk in enumerate(product_id_chunks):
        logging.info(f"Processing chunk {index + 1}/{len(product_id_chunks)}...")
        template_wb = load_workbook(template_path)

        # 写入“Warehouse stock”表单
        worksheet_warehouse = template_wb["仓库库存"]

        # 确定仓库名称
        warehouse_name = ck_name[0] if num == 0 else ck_name[index]

        # 确定开始行
        row_to_start = 2
        
        # 写入整行数据
        for product_index, product_id in enumerate(chunk):
            row_number = product_index + row_to_start
            row_data = (warehouse_name, product_id, '', int(stock_count))
            for col, value in enumerate(row_data, start=1):
                worksheet_warehouse.cell(row=row_number, column=col, value=value)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_库存_' + str(index + 1) + '_' + symbol + ".xlsx"

        # 保存当前工作簿
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")

    zipFile.close()
    logging.info(f"Zip file created: {zipName}")

    return zipName

# 多配色挂毯
def generateGuatanWithColor(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Length of the Larger Side, cm'] = length_of_sku[j]
            df.loc[i, 'Length of the smaller side, cm'] = width_of_sku[j]
            df.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
    zipFile.close()
    return zipName

# 多配色挂毯
def generateGuatanWithColor2(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Length of the Larger Side, cm*'] = length_of_sku[j]
            df.loc[i, 'Length of the smaller side, cm'] = width_of_sku[j]
            df.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
    zipFile.close()
    return zipName


#多配色挂毯、浴帘生成库存
def generateStockGTWithColor(input, name, num, id_symbol, ck_name, stock_count, files_symbol):
    logging.info("Starting generateStock function...")

    newNameLists = []
    template_path = './generateAndsplit/stock-update-template-cn.xlsx'
    zipName = name.split(".")[0] + "zip_库存.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')

    logging.info("Reading Excel file...")
    df = pd.read_excel(input, sheet_name="Template", header=1)

    # 生成id列表
    product_id_list = []
    logging.info("Generating product ID list...")
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            product_id = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            product_id_list.append(product_id)

    # 当num为0时，不分割product_id_list
    if num == 0:
        product_id_chunks = [product_id_list]
    else:
        product_id_chunks = [product_id_list[i:i + num] for i in range(0, len(product_id_list), num)]

    # 检查分割后数组的长度与ck_name的长度是否一致
    if len(product_id_chunks) != len(ck_name) and num != 0:
        error_message = 'product_id_list的分割长度与ck_name的长度不一致'
        logging.error(error_message)
        return {'error': error_message}

    # 写入数据
    logging.info("Writing data...")
    for index, chunk in enumerate(product_id_chunks):
        logging.info(f"Processing chunk {index + 1}/{len(product_id_chunks)}...")
        template_wb = load_workbook(template_path)

        # 写入“Warehouse stock”表单
        worksheet_warehouse = template_wb["仓库库存"]

        # 确定仓库名称
        warehouse_name = ck_name[0] if num == 0 else ck_name[index]

        # 确定开始行
        row_to_start = 2
        
        # 写入整行数据
        for product_index, product_id in enumerate(chunk):
            row_number = product_index + row_to_start
            row_data = (warehouse_name, product_id, '', int(stock_count))
            for col, value in enumerate(row_data, start=1):
                worksheet_warehouse.cell(row=row_number, column=col, value=value)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_库存_' + str(index + 1) + '_' + symbol + ".xlsx"

        # 保存当前工作簿
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        logging.info(f"Saved file: {newName}")

    zipFile.close()
    logging.info(f"Zip file created: {zipName}")

    return zipName

# 多配色浴帘
def generateYulianWithColor(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Height, cm*'] = length_of_sku[j]
            df.loc[i, 'Width, cm'] = width_of_sku[j]
            df.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
    zipFile.close()
    return zipName


# 多配色三件套
def generateSanjiantaoWithColor(input, name, num, id_symbol, price, price_before_discount, length_of_sku, width_of_sku, files_symbol):
    newNameLists = []

    #复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Template", "Ozone.Video cover", "Ozone.Video"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第4行到最后一行
            if max_row >= 4:
                worksheet.delete_rows(4, max_row - 3)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Template",header=1)


    # 视频封面直接获取第一个，之后所有的sku都设置一样的视频
    try:
        df_video_cover = pd.read_excel(input, sheet_name="Ozone.Video cover", header=1)
        if df_video_cover.shape[0] > 1:
            video_cover_URL = df_video_cover.iloc[1]["Ozone.Video Cover: URL"] if pd.notna(df_video_cover.iloc[1]["Ozone.Video Cover: URL"]) else ''
            print(f"video_cover_URL: {video_cover_URL}")
        else:
            video_cover_URL = ''
            print("No video cover data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_cover_URL = ''
        print("Error reading video cover details. Skipping...")
    
    # 视频直接获取第一行，之后所有的sku都设置一样的视频
    try:
        df_video = pd.read_excel(input, sheet_name="Ozone.Video", header=1)
        if df_video.shape[0] > 1:
            video_name = df_video.iloc[1]["Ozone.Video: Name"] if pd.notna(df_video.iloc[1]["Ozone.Video: Name"]) else ''
            video_url = df_video.iloc[1]["Ozon.Video: URL"] if pd.notna(df_video.iloc[1]["Ozon.Video: URL"]) else ''
            video_products_on_video = df_video.iloc[1]["Ozone.Video: products on video"] if pd.notna(df_video.iloc[1]["Ozone.Video: products on video"]) else ''
            print(f"video_name: {video_name}")
            print(f"video_url: {video_url}")
            print(f"video_products_on_video: {video_products_on_video}")
        else:
            video_name = video_url = video_products_on_video = ''
            print("No video data found in the first row. Skipping...")
    except (IndexError, KeyError):
        video_name = video_url = video_products_on_video = ''
        print("Error reading video details. Skipping...")

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    new_video_cover= []
    new_video= []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(len(id_symbol)):
            df.loc[i, 'Price, USD*'] = price[j]
            df.loc[i, 'Price before discount, USD'] = price_before_discount[j]
            df.loc[i, 'Duvet Cover Size'] = length_of_sku[j]
            df.loc[i, 'Bed Sheet Size'] = width_of_sku[j]
            df.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)

            df_video_cover.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video_cover.loc[i, "Ozone.Video Cover: URL"] = video_cover_URL if video_cover_URL else ''
            row = df_video_cover.iloc[i].tolist()
            new_video_cover.append(row)

            df_video.loc[i, "Article code*"] = df.iloc[i]["Model Name (to combine products into one PDP)*"] + df.iloc[i]["Product color"] + id_symbol[j]
            df_video.loc[i, "Ozone.Video: Name"] = video_name if video_name else ''
            df_video.loc[i, "Ozon.Video: URL"] = video_url if video_url else ''
            df_video.loc[i, "Ozone.Video: products on video"] = video_products_on_video if video_products_on_video else ''
            row = df_video.iloc[i].tolist()
            new_video.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    video_cover_columns = df_video_cover.columns.tolist()
    video_columns = df_video.columns.tolist()

    new_df = pd.DataFrame(new_rows, columns=columns)
    new_video_cover_df = pd.DataFrame(new_video_cover, columns=video_cover_columns)
    new_video_df = pd.DataFrame(new_video, columns=video_columns)

    # 写入数据
    for index, (li_supplier, li_video_cover, li_video) in enumerate(zip(cut(new_df, num), cut(new_video_cover_df, num), cut(new_video_df, num))):
        template_wb = load_workbook(template_path)
        
        # Write to "Template" sheet
        worksheet_supplier = template_wb["Template"]

        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        # 添加文件名标记
        symbol = files_symbol[index] if files_symbol and len(files_symbol) > index else ""
        newName = name.split(".")[0] + '_商品_' + str(index + 1) + '_' + symbol + ".xlsx"

        for row in dataframe_to_rows(li_supplier, index=False, header=False):
            worksheet_supplier.append(row)

        # 如果 video_cover_URL 不为空，则写入 "Ozone.Video cover" sheet
        if video_cover_URL:
            worksheet_video_cover = template_wb["Ozone.Video cover"]
            for row in dataframe_to_rows(li_video_cover, index=False, header=False):
                worksheet_video_cover.append(row)

        # 如果 video_name, video_url, video_products_on_video 不为空，则写入 "Ozone.Video" sheet
        if video_name or video_url or video_products_on_video:
            worksheet_video = template_wb["Ozone.Video"]
            for row in dataframe_to_rows(li_video, index=False, header=False):
                worksheet_video.append(row)
        
        logging.info(f"files_symbol{files_symbol}")
        if files_symbol and len(files_symbol) >1 :
            logging.info(f"files_symbolindex{files_symbol[index]}")
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        logging.info(f"Saved file: {newName}")
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
    zipFile.close()
    return zipName


# TikTok服饰通用版
def generateTikTokCloth(input, name, manufacturer_sizes):
    newNameLists = []

    # 复制输入的文件，并且删除多余的文件内容作为输出的模板
    template_input_path = os.path.join(os.path.dirname(input), 'template_' + os.path.basename(input))
    shutil.copy(input, template_input_path)

    template_wb = load_workbook(template_input_path)

    sheets_to_clear = ["Sheet1"]
    for sheet_name in sheets_to_clear:
        try:
            worksheet = template_wb[sheet_name]
            max_row = worksheet.max_row
            # 删除第2行到最后一行
            if max_row > 1:
                worksheet.delete_rows(2, max_row)
        except KeyError:
            print(f"Sheet {sheet_name} not found in workbook. Skipping...")
            continue

    template_wb.save(template_input_path)
    template_path = template_input_path

    zipName = name.split(".")[0] + "_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    df = pd.read_excel(input, sheet_name="Sheet1", header=0)

    # 将每一行按照选择的尺码生成，并将它们在一起
    new_rows = []
    for i in range(len(df)):
        logging.info(f"i: {i}")
        logging.info(f"len(df): {len(df)}")
        for j in range(len(manufacturer_sizes)):
            # 复制一份原始行数据
            new_row = df.iloc[i].copy()
            # 更新变种属性值二和变种名称
            new_row['变种属性值二'] = manufacturer_sizes[j]
            new_row['变种名称'] = f'black，{manufacturer_sizes[j]}'
            #logging.info(f"new_row): {new_row}")
            # 将新行添加到列表中
            new_rows.append(new_row)

    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)

    # 写入数据
    template_wb = load_workbook(template_path)

    # Write to "Sheet1" sheet
    worksheet_supplier = template_wb["Sheet1"]

    # 添加文件名标记
    newName = name.split(".")[0] + '_商品.xlsx'

    for row in dataframe_to_rows(new_df, index=False, header=False):
        worksheet_supplier.append(row)

    template_wb.save("./static/{}".format(newName))
    newNameLists.append(newName)
    logging.info(f"Saved file: {newName}")
    # zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)

    # zipFile.close()
    return newName



# 多配色T恤
def generateMultiColor(input,name,num):
    newNameLists = []
    template_path = './generateAndsplit/mb_Supplier template.xlsx'
    zipName = name.split(".")[0]+"zip_商品.zip"
    zipFile = zipfile.ZipFile("./static/{}".format(zipName), 'w')
    # df = pd.read_excel('input.xlsx', sheet_name="Vendor template",header=1)
    df = pd.read_excel(input, sheet_name="Supplier template",header=1)
    russian_sizes = ["46", "48", "50", "52", "54", "56"]
    manufacturer_sizes = ['L', 'XL', '2XL', '3XL', '4XL', '5XL']
    # 将每一行分成六份，并将它们在一起
    new_rows = []
    for i in range(len(df)):
        if i == 0:
            continue
        for j in range(6):
            df.loc[i, 'Manufacturer Size'] = manufacturer_sizes[j]
            df.loc[i, 'Russian Size*'] = russian_sizes[j]
            df.loc[i, 'Size of the Product in the Photo'] = russian_sizes[j]
            if i>=2:
                if df.loc[i, "Product name"] == df.loc[i-1, "Product name"]:
                    df.loc[i, "Merge on One PDP*"] = df.iloc[i-1]["Merge on One PDP*"]
                else:
                    buff =  str(int(df.iloc[i-1]["Merge on One PDP*"][-3:])+1)
                    if len(buff) == 2:
                        buff = "0"+buff
                    elif len(buff) == 1:
                        buff = "00"+buff
                    df.loc[i, "Merge on One PDP*"] = df.iloc[i-1]["Merge on One PDP*"][:-3] +  buff
            #print(i,df.iloc[i]["Merge on One PDP*"],df.iloc[i]["Product color*"],manufacturer_sizes[j])
            df.loc[i, "Product ID*"] = df.iloc[i]["Merge on One PDP*"] + " " + df.iloc[i]["Product color*"] + " " + manufacturer_sizes[j]
            row = df.iloc[i].tolist()
            new_rows.append(row)
    columns = df.columns.tolist()
    new_df = pd.DataFrame(new_rows, columns=columns)
    # 写入数据
    for index,li in enumerate(cut(new_df,num)):
        template_wb = load_workbook(template_path)
        worksheet = template_wb.active
        newName = name.split(".")[0] + '_'+ str(index+1) +"_商品.xlsx"
        for row in dataframe_to_rows(li, index=False, header=False):
            worksheet.append(row)
        for cell in worksheet['U']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        for cell in worksheet['AF']:
            if cell.value is not None:
                try:
                    cell.value = int(cell.value)
                except:
                    pass
        template_wb.save("./static/{}".format(newName))
        newNameLists.append(newName)
        zipFile.write("./static/{}".format(newName), newName, zipfile.ZIP_DEFLATED)
        # new_df.to_excel("../static/{}".format(newName), index=False,header=None)
    zipFile.close()
    return zipName

# n = bq("input.xlsx","input.xlsx",7)
# print(n)